# import
import glob,os
import pandas as pd
from openpyxl import load_workbook
# import -- End


class CustomException(Exception):
    """ my custom exception class """

class MergeAllExcelFiies:
    # functions
    def get_sheetnames_xlsx(filepath):
        wb = load_workbook(filepath, read_only=True, keep_links=False)
        return wb.sheetnames

    os.system('cls')
    os.chdir('c:\\FranchiseData\\Data\\')
    # print('in clear block...........')
    try:
        os.remove('AllDataMerged.csv')
        os.remove('logList.txt')    
    except FileNotFoundError:
        pass    
    
    dataFiles=[]
    df=pd.DataFrame()
    df1=pd.DataFrame()
    file_list = glob.glob('*.*')
    logList=[]

    for file in file_list:
        if file.endswith('.xlsx') or file.endswith('.xls') :
            try:
                print('block1')
                sheetNames=[]
                sheetNames=get_sheetnames_xlsx(file)
                sheet=sheetNames[0]
                
                if 'SALE' in sheet.upper():
                    print(sheet)
                    sheetName=sheet                
                    dataFiles=pd.read_excel(file,sheet_name=f'{sheetName}',index_col=False
                                            ,header=None,skiprows=[0]
                                            )
                    totalCol=len(dataFiles.columns)
                    # print(dataFiles.info())

                    if totalCol==26:                    
                        # print(f"processing file {file},file, total columns length {totalCol}")
                        totalRows=len(dataFiles)
                        logList.append('\n')
                        logList.append(f'processing file :+{file}, total rows {totalRows}') 
                                   
                        df=pd.concat([dataFiles])
                else:
                    print(f"else {sheet}")
                    raise CustomException(f'Invalid sheet name found in {file}')

            except CustomException as ex:
                print(ex)   
                logList.append('\n')
                logList.append(str(ex)) 
                pass
            except Exception as ex:
                exError=f'error in a file {file}: {ex}'
                # print(exError)
                logList.append('\n')
                logList.append(str(exError))        
                pass
            finally:
                print('finally...')
                totalCol = 0

    print('df size ',len(df))
    file = open('logList.txt', 'w')
    file.writelines(logList)
    file.close()
    
    # print('df col')
    if len(df)>0:
        df.columns = ('Franchise Customer OrderNo', 'Franchise Customer Invoice Date',
                    'Franchise Customer Invoice Number','Channel','Franchise Code',
                    'Franchise Customer Number','IBL Customer Number','RD Customer Name',
                    'IBL Customer Name',      'Customer Address',        'Franchise Item Code',
                    'IBL Item Code',     'Franchise Item Description',    'IBL Item Description',
                    'Quantity Sold',    'Gross Amount',          'Reason',           'FOC',         'BATCH_NO',
                    'PRICE',        'BON_QTY',                 'DISC_AMT',             'NET_AMT',
                    'DISCOUNTED_RATE',               'Brick code',                  'Brick Name',
                    )

        df['Franchise Customer Invoice Date'] = pd.to_datetime(
            df['Franchise Customer Invoice Date'])
        
        df.to_csv('AllDataMerged.csv'
                ,index=False
                #   ,header=None
                )

    print('Process completed, check logList.txt file')

